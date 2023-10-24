# encoding=utf-8
import argparse
import logging
import os
import re
import unicodedata
from dataclasses import dataclass
from functools import wraps
import hashlib
from pathlib import Path
from time import time
from typing import Generator
from itertools import groupby
from tqdm import tqdm

from dotenv import load_dotenv
import pandas as pd
import orjson
from elasticsearch import Elasticsearch, helpers
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from bibbi.console.config import Config
from bibbi.db import Db, ColumnDataTypes
from bibbi.promus_service import PromusService

log = logging.getLogger(__name__)


def timing(f):
    @wraps(f)
    def wrap(*args, **kw):
        ts = time()
        result = f(*args, **kw)
        te = time()
        log.info('%r took: %2.4f sec', f.__name__, te - ts)
        return result
    return wrap


@dataclass
class AuthorityTable:
    name: str
    primary_key: str
    field: str
    approve_column: str = 'Approved'
    title_column: str = 'NULL'
    display_column: str = '_DisplayValue'


@dataclass
class DataRow:
    id: str
    authorities: list


@dataclass
class ReportHeader(object):
    line1: str
    line2: str
    width: int


class Cache:
    def __init__(self, dir, max_age,):
        self.dir = Path(dir)
        self.max_age = max_age

    def get(self, key):
        cache_file = self.dir.joinpath(key)
        if not cache_file.exists():
            return None
        cache_age = time() - cache_file.stat().st_mtime
        if cache_age > self.max_age:
            log.info('Cache age: %d secs => Too old, will refresh.', cache_age)
            return None
        log.info('Cache age: %d secs => Using cache.', cache_age)
        return pd.read_pickle(str(cache_file))

    def put(self, key, data):
        cache_file = self.dir.joinpath(key)
        cache_file.parent.mkdir(exist_ok=True)
        data.to_pickle(str(cache_file))


class Report:

    def __init__(self, headers):
        self.headers = headers
        self.data = []

    @staticmethod
    def remove_control_characters(s):
        return ''.join(ch for ch in s if unicodedata.category(ch)[0] != 'C')

    def add(self, row, wash=False):
        # Jeg fant bare én post med kontrolltegn i hele basen, så sannsynligheten for å støte på dem er så
        # lav at jeg deaktiverer vasking inntil videre, selv om det ikke krever noe særlig ekstra behandlingstid.
        if len(row) != len(self.headers):
            raise ValueError('Invalid row (1): %s' % str(row))
        if wash:
            washed = [self.remove_control_characters(col) for col in row]
            if washed != row:
                log.warning('Row included control characters: %s', str(row))
            self.data.append(washed)
        else:
            self.data.append(row)

    def save(self, filename):
        wb = Workbook()
        ws = wb.active

        for n, header in enumerate(self.headers):
            ws.column_dimensions[chr(65 + n)].width = header.width

        ws.freeze_panes = 'A3'

        header_font = Font(bold=True)
        header_fill = PatternFill('solid', fgColor='FFFFEE')
        link_font = Font(color='0000FF')

        for n, header in enumerate(self.headers):
            cell = ws.cell(row=1, column=n + 1, value=header.line1)
            cell.font = header_font
            cell.fill = header_fill

            cell = ws.cell(row=2, column=n + 1, value=header.line2)
            cell.font = header_font
            cell.fill = header_fill

        for row_no, values in enumerate(self.data):
            for col_no, value in enumerate(values):
                link = None
                if value.startswith('{BIBBI}'):
                    value = value[7:]
                    link = 'https://id.bs.no/bibbi/' + value
                elif value.startswith('{NORAF}'):
                    value = value[7:]
                    link = 'https://bsaut.toolforge.org/show/' + value

                cell = ws.cell(row=row_no + 3, column=col_no + 1, value=value)

                if link is not None:
                    cell.hyperlink = link
                    cell.font = link_font

        last_row_no = len(self.data) + 2
        last_col_chr = chr(64 + len(self.headers))
        ws.auto_filter.ref = "A2:%s%s" % (last_col_chr, last_row_no)

        wb.save(filename)
        log.info('Wrote %d data rows to %s', len(self.data), filename)


class Runner:

    def __init__(self, promus_adapter: PromusService, cache: Cache, config: Config):
        # self.es = elasticsearch
        self.conn = promus_adapter.connection
        self.cache = cache
        self.config = config

    def get_data(self, query: str, params: ColumnDataTypes = None, **kwargs) -> pd.DataFrame:
        key = hashlib.sha1(query.encode('utf-8')).hexdigest()
        df = self.cache.get(key)
        if df is None:
            df = self.conn.select_dataframe_sa(query, params, **kwargs)
            self.cache.put(key, df)
        return df

    @timing
    def get_authority_links(self):
        return self.get_data('''
            SELECT
                ItemField_ID as field_id,
                LTRIM(STR(Item_ID)) AS item_id,
                LTRIM(STR(Authority_ID)) AS local_id,
                FieldCode as field
            FROM ItemField
            WHERE
                Authority_ID IS NOT NULL AND Authority_ID != '0'
                AND (
                    FieldCode LIKE '1%'
                    OR FieldCode LIKE '6%'
                    OR FieldCode LIKE '7%'
                    OR FieldCode LIKE '8%'
                )
        ''')
        # Analytiske biinførsler: Hadde først utelatt dem med Indicator2 <> 2, men husker ikke hvorfor. Det fører til
        # at vi får noen autoriteter som tilsynelatende ikke er i bruk i Skosmos, som https://id.bs.no/bibbi/1023449,
        # så prøver å ta dem med og se hvordan det går. DM 2021-08-04

    @timing
    def get_authority_table(self, table: AuthorityTable) -> pd.DataFrame:
        # print(table)
        return self.get_data('''
            SELECT
                LTRIM(STR(%(primary_key)s)) AS local_id,
                LTRIM(STR(Bibsent_ID)) as bibsent_id,
                NotInUse as not_in_use,
                %(approve_column)s AS approved,
                %(title_column)s AS title,
                %(display_column)s AS label
            FROM %(name)s
        ''' % table.__dict__)

    @timing
    def get_authorities(self):
        return {
            table.field: self.get_authority_table(table) for table in [
                AuthorityTable('AuthorityWork', 'WorkID', '01', title_column='TitlePreferred'),
                AuthorityTable('AuthorityPerson', 'PersonID', '00', title_column='TopicTitle'),
                AuthorityTable('AuthorityCorp', 'CorpID', '10', title_column='TopicTitle'),
                AuthorityTable('AuthorityConf', 'ConfID', '11'),
                # AuthorityTable('AuthorityTitle', 'TitleID', '30'),      # Har ikke Bibsent_ID
                AuthorityTable('AuthorityTopic', 'AuthID', '50'),
                AuthorityTable('AuthorityGeoTopic', 'GeoTopicID', '51', display_column='_GeoName'),
                AuthorityTable('AuthorityFreeGenre', 'TopicID', '53', "CAST(1 AS BIT)"),  # Mangler 'Approved'-felt
                AuthorityTable('AuthorityGenre', 'TopicID', '55', "CAST(1 AS BIT)"),
            ]
        }

    @timing
    def get_marc_data(self):
        return self.get_data('''
          SELECT
            LTRIM(STR(ItemField.Item_ID)) AS item_id,
            ItemField.FieldCode AS field,
            ItemSubField.SubFieldCode AS subfield,
            ItemSubField.Text AS value
          FROM ItemField
          INNER JOIN ItemSubField ON ItemSubField.ItemField_ID = ItemField.ItemField_ID
          WHERE ItemField.FieldCode IN ('019', '020', '024', '082', '245', '260')
          ORDER BY ItemField.Item_ID, ItemField.FieldCode
        ''')

    @timing
    def get_items(self):
        return self.get_data('''
            SELECT
                LTRIM(STR(Item_ID)) AS item_id,
                ApproveDateFirst AS cataloguing_date,
                ApproveDate AS approve_date,
                ItemYear as pub_year,
                Title AS title_ax,
                --, dbo.fn_ItemSubFieldText(Item_ID, '245', 'a') as f245a,
                --, dbo.fn_ItemSubFieldText(Item_ID, '245', 'b') as f245b,
                --, dbo.fn_ItemSubFieldText(Item_ID, '260', 'c') as f260c,
                Varenr AS varenr,
                LTRIM(STR(BibbiNr)) AS bibbi_id
            FROM Item
            WHERE ApproveDateFirst IS NOT NULL
            AND ApproveDate IS NOT NULL
            -- AND ApproveDate >= '2019'
        ''', dont_touch=['cataloguing_date', 'approve_date'], )

    @timing
    def get_export2ax(self):
        return self.get_data('''
            SELECT
                ID AS id,
                LTRIM(STR(PromusID)) AS item_id,
                EAN AS ean,
                DocumentType AS doc_type
            FROM Export_PromusToAx
        ''')


    @timing
    def get_person_roles(self):
        res = self.get_data('''
            SELECT
                ItemField_ID AS field_id,
                Text AS value
            FROM ItemSubField
            WHERE SubField_ID = 33   -- 100 $e
               OR SubField_ID = 219  -- 700 $e
        ''')
        res.value = res.value.str.lower()
        return res

    @timing
    def get_doc_types(self):
        res = self.get_data('''
            SELECT
                DocFieldName AS code,
                DocText AS value,
                RDAcontent AS rda_content,
                RDAmedia AS rda_media,
                RDAcarrier AS rda_carrier
            FROM EnumDocTypes
        ''')
        return res

    def convert(self, items, export2ax, marc_data, roles, doc_types, authority_links, authorities) -> Generator:
        valid = 0
        invalid = 0

        report = Report([
            ReportHeader('Vare', 'ID', 12),
            ReportHeader('', 'Tittel', 60),
            ReportHeader('', 'Godkjent', 12),
            ReportHeader('Autoritetskobling', 'MARC-felt', 10),
            ReportHeader('', 'Feil', 30),
            ReportHeader('Autoritet', 'ID', 12),
            ReportHeader('', 'Streng', 80),
        ])

        #print("SORT")
        # export2ax = sorted(export2ax, key=lambda x: x.id, reverse=True)
        export2ax_map = {}
        for rec in tqdm(export2ax.itertuples()):
            cur = export2ax_map.get(rec.item_id)
            if cur is None or rec.id > cur.id:
                export2ax_map[rec.item_id] = rec
        del export2ax # Free some memory (maybe)
        log.info('✔ export2ax map done (%d items)', len(export2ax_map))

        # Map: item.item_id -> MARC fields
        marc_map = {}
        for rec in tqdm(marc_data.itertuples()):
            if rec.item_id not in marc_map:
                marc_map[rec.item_id] = {}
            k = rec.field + '$' + rec.subfield
            marc_map[rec.item_id][k] = rec.value  # Note: If multiple values, only the last one is used
        del marc_data  # Free some memory (maybe)
        log.info('✔ MARC map done (%d items)', len(marc_map))

        # Map: aut.local_id -> aut.bibsent_id
        authority_map = {}
        for key, df in authorities.items():
            for authority in df.itertuples():
                authority_map['%s-%s' % (key, authority.local_id)] = authority

        # Map: field.field_id -> role
        role_map = {}
        for rec in tqdm(roles.itertuples()):
            role_map[rec.field_id] = rec.value

        # Map: doc_type.code -> value
        doctype_map = {}
        for rec in tqdm(doc_types.itertuples()):
            doctype_map[rec.code] = rec.value


        # Map: item.bibbi_id -> [authority.bibsent_id]
        item_map = {}
        for link in tqdm(authority_links.itertuples()):
            code = link.field[1:]
            sig = '%s-%s' % (code, link.local_id)
            aut_role = role_map.get(link.field_id)
            try:
                authority = authority_map[sig]
                item_map[link.item_id] = item_map.get(link.item_id, []) + [(link.field, authority, aut_role)]
                valid += 1
            except KeyError:
                if code not in authorities:
                    log.debug('Cannot authorize field %s', link.field)
                else:
                    invalid += 1
                    log.warning('Item %s contains link from field X%s to unknown authority %s',
                                link.item_id, link.field, link.local_id)

        log.info('Authority links: %d of %d invalid (%.2f %%)',
                 invalid, valid + invalid, (invalid / (valid + invalid) * 100))

        def warn(item, marc_field, authority, msg):
            #log.warning('Katalogpost %s ("%s", godkjent %s) har et %s-felt %s: %s (%s)',
            #            item.item_id, item.title_ax, item.cataloguing_date.strftime('%Y-%m-%d'), marc_field, msg, authority.local_id, authority.label)
            report.add([item.item_id, item.title_ax, item.cataloguing_date.strftime('%Y-%m-%d'), marc_field, msg, authority.local_id, authority.label])

        record_count = 0
        link_count = 0
        error_counts = {
            'NotInUse-autoritet': 0,
            'ikke-godkjent autoritet': 0,
            'mangler Bibbi-ID': 0,
        }
        for item in tqdm(items.itertuples()):
            # print(item.item_id)
            export2ax_item = export2ax_map.get(item.item_id)
            marc_map_item = marc_map.get(item.item_id, {})
            title = marc_map_item.get('245$a')
            subtitle = marc_map_item.get('245$b')
            if subtitle is not None:
                title += ' : ' + subtitle
            doc = {
                '_id': item.bibbi_id,
                # 'title_ax': item.title_ax,
                'ean': marc_map_item.get('025$a') or marc_map_item.get('020$a'),
                'webdewey': marc_map_item.get('082$a'),
                'cataloguing_date': item.cataloguing_date.strftime('%Y-%m-%d'),
                'approve_date': item.approve_date.strftime('%Y-%m-%d'),
                'title': title,
                'pub_year': item.pub_year,
                'pub_place': marc_map_item.get('260$a'),
                'publisher': marc_map_item.get('260$b'),
                'authorities': [],
                'doc_types': marc_map_item.get('019$b'),
                'form': marc_map_item.get('019$d'),
            }
            if doc['webdewey'] is not None:
                doc['webdewey'] = doc['webdewey'].replace('/', '')
            if export2ax_item:
                # doc['ean'] = export2ax_item.ean
                doc['doc_type'] = export2ax_item.doc_type
            if doc['doc_types'] is None:
                doc['doc_types'] = []
            else:
                doc['doc_types'] = [doctype_map.get(x, x) for x in doc['doc_types'].split(',')]

            form_map = {
                "A": "Antologi",
                "B": "Billedbok",
                "D": "Dikt",
                "L": "Lærebok",
                "N": "Novelle",
                "P": "Pekebok",
                "R": "Roman",
                "S": "Skuespill",
                "T": "Tegneserie",
            }
            doc['form'] = form_map.get(doc['form'], doc['form'])

            for marc_field, authority, aut_role in item_map.get(item.item_id, []):
                link_count += 1
                if authority.not_in_use == 'True':
                    warn(item, marc_field, authority, 'NotInUse-autoritet')
                    error_counts['NotInUse-autoritet'] += 1

                elif authority.approved == 'False':
                    warn(item, marc_field, authority, 'ikke-godkjent autoritet')
                    error_counts['ikke-godkjent autoritet'] += 1
                    lab = authority.label.replace('&', '').replace('-', '').replace(':', '').lower().split(' ')

                elif pd.isnull(authority.bibsent_id):
                    warn(item, marc_field, authority, 'mangler Bibbi-ID')
                    error_counts['mangler Bibbi-ID'] += 1

                else:
                    authority_dict = {
                        'id': authority.bibsent_id,
                        'label': authority.label,
                        # Obs: label er greit som søkehjelp, men ikke alltid visning.
                        # Eks: Smaaland, Tor : 1958- : n.  : Småland, Tor
                        'type': 'unknown',
                    }
                    if aut_role is not None:
                        authority_dict['role'] = aut_role
                    if marc_field.startswith('1') or marc_field.startswith('7'):
                        if authority.title:
                            authority_dict['type'] = 'work'
                            authority_dict['title'] = authority.title
                        else:
                            authority_dict['type'] = 'creator'
                            if marc_field.startswith('10') or marc_field.startswith('70'):
                                authority_dict['is_person'] = True

                    elif marc_field.startswith('655'):
                        authority_dict['type'] = 'genre'
                    elif marc_field.startswith('651'):
                        # Dog... Oslo - Grønland - Kulturhistorie er ikke et sted..
                        authority_dict['type'] = 'place'
                    elif marc_field.startswith('6'):
                        authority_dict['type'] = 'topic'
                    #else:
                    #    # log.debug('%s - Ignoring %s field', item.item_id, marc_field)
                    #    pass
                    doc['authorities'].append(authority_dict)

            record_count += 1
            doc['_id'] = item.bibbi_id
            yield doc

        log.info('Behandlet %d katalogposter med %d autoritetskoblinger', record_count, link_count)
        for error, cnt in error_counts.items():
            log.info(' - %d feil av typen "%s" (%.2f %%)', cnt, error, cnt / link_count * 100)

        #log.info('%d terms. Top terms:', len(terms))
        #groups = [(key, len(list(group))) for key, group in groupby(sorted(terms))]
        #for top_group in sorted(groups, key=lambda x: x[1], reverse=True)[:10]:
        #    print(top_group)

        total_errors = sum(error_counts.values())
        log.info('Totalt %d lenker (%.2f %%) som ikke peker til en godkjent autoritetspost',
                 total_errors, total_errors / link_count * 100)

        report.save('ugyldige_autoritetskoblinger.xlsx')

    @timing
    def dump(self, data) -> bytes:
        return orjson.dumps(data, option=orjson.OPT_INDENT_2 | orjson.OPT_APPEND_NEWLINE)

    @timing
    def push_to_es(self, records):
        try:
            # make the bulk call, and get a response
            response = helpers.bulk(self.es, records, index='bibbi_cat', doc_type='bib_record')
            print("\nRESPONSE:", response)
        except Exception as e:
            print("\nERROR:", e)

    def run(self):
        log.info('Get export2ax')
        export2ax = self.get_export2ax()
        log.info("Got %d rows", len(export2ax))

        log.info('Get items')
        items = self.get_items()

        # Might be faster
        log.info('Get marc data')
        marc_data = self.get_marc_data()
        roles = self.get_person_roles()
        doc_types = self.get_doc_types()

        log.info('Get links')
        authority_links = self.get_authority_links()
        authorities = self.get_authorities()
        log.info('Got %d items, %d authority_links, %d authorities',
                 len(items), len(authority_links), len(authorities))

        dest_file = self.config.dest_dir.joinpath('catalog.jsonl')

        nrecs = 0
        with dest_file.open('wb') as fp:
            for record in self.convert(items, export2ax, marc_data, roles, doc_types, authority_links, authorities):
                fp.write(orjson.dumps(record, option=orjson.OPT_APPEND_NEWLINE))
                nrecs += 1

        log.info('Wrote %d records to %s', nrecs, dest_file.name)

        # for record in records[:20]:
        #     print(orjson.dumps(record, option=orjson.OPT_INDENT_2 | orjson.OPT_APPEND_NEWLINE).decode('utf-8'))

        # return

        #with self.config.dest_dir.joinpath('catalog.json').open('wb') as fp:
        #    fp.write(self.dump(records))

        #self.push_to_es(records)



def get_services(config: Config):
    promus_adapter = PromusService(connection=Db(**{
        'server': os.getenv('DB_SERVER'),
        'port': os.getenv('DB_PORT'),
        'database': os.getenv('DB_DB'),
        'user': os.getenv('DB_USER'),
        'password': os.getenv('DB_PASSWORD'),
    }))

    #elasticsearch_adapter = Elasticsearch(
    #    [os.getenv('ELASTICSEARCH_HOST')],
    #    http_auth=(os.getenv('ELASTICSEARCH_USER'), os.getenv('ELASTICSEARCH_PASSWORD'))
    #)

    return {
        'config': config,
        'promus_adapter': promus_adapter,
        #'elasticsearch': elasticsearch_adapter,
        'cache': Cache('cache', 36000),
    }


def extract_catalog(config: Config, options):

    # configure_logging()

    services = get_services(config)
    if options.no_cache:
        services['cache'].max_age = 0

    Runner(**services).run()
